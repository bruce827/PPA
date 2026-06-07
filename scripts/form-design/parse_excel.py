#!/usr/bin/env python3
"""
Excel 表单模板解析器
将 Excel 表单模板转换为 JSON 格式，用于 Web 预览页面
"""

import json
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 列映射定义（基于 excel填写规范.md）
COLUMN_MAPPING = {
    # 表单信息
    '应用名称': 'app_name',
    '表单名称': 'form_name',
    '表单编码': 'form_code',
    '过滤条件': 'filter_condition',

    # 基本信息-数据库信息
    '字段名称': 'field_name',
    '字段编码': 'field_code',
    '是否主键': 'is_primary_key',
    '是否虚拟字段': 'is_virtual',
    '字段类型': 'field_type',
    '字段长度': 'field_length',
    '字段精度': 'field_precision',
    '默认值': 'default_value',

    # 基本信息
    '输入类型': 'input_type',
    '输入参数': 'input_params',
    '是否必填': 'is_required',
    '是否唯一': 'is_unique',
    '提示信息': 'placeholder',
    '备注': 'remark',

    # 卡片页面信息
    '卡片分组': 'card_group',
    '卡片排序': 'card_sort',
    '卡片宽度': 'card_width',
    '新增控制': 'add_control',
    '更新控制': 'update_control',
    '详情控制': 'detail_control',

    # 列表页面信息
    '列表宽度': 'list_width',
    '列表控制': 'list_control',
    '列表排序': 'list_sort',
    '列表格式化': 'list_formatter',

    # 过滤条件
    '是否过滤器': 'is_filter',
    '过滤方式': 'filter_mode',
    '过滤器默认值': 'filter_default',
    '过滤器提示信息': 'filter_placeholder',
    '来源系统': 'source_system',
}

# 输入类型映射（Excel 编码 → 前端组件）
INPUT_TYPE_MAPPING = {
    'text': {'component': 'Input', 'label': '文本框'},
    'textArea': {'component': 'Input.TextArea', 'label': '文本域'},
    'number': {'component': 'InputNumber', 'label': '数字框'},
    'selectorDropdown': {'component': 'Select', 'label': '下拉选择器'},
    'selectorDropdownTree': {'component': 'TreeSelect', 'label': '下拉树形选择器'},
    'selectorPopup': {'component': 'Modal', 'label': '弹出选择器'},
    'selectorPopupInput': {'component': 'AutoComplete', 'label': '弹出选择器(可输入)'},
    'selectorDate': {'component': 'DatePicker', 'label': '日期选择器'},
    'selectorTime': {'component': 'TimePicker', 'label': '时间选择器'},
    'selectorDateInterval': {'component': 'DatePicker.RangePicker', 'label': '日期区间'},
    'selectorTimeInterval': {'component': 'TimePicker.RangePicker', 'label': '时间区间'},
    'selectorYear': {'component': 'DatePicker', 'label': '年份选择器', 'props': {'mode': 'year'}},
    'selectorYearMonth': {'component': 'DatePicker', 'label': '年月选择器', 'props': {'mode': 'month'}},
    'radioBox': {'component': 'Radio.Group', 'label': '单选框'},
    'checkBox': {'component': 'Checkbox.Group', 'label': '复选框'},
    'uploadImage': {'component': 'Upload', 'label': '图片上传'},
    'uploadFile': {'component': 'Upload', 'label': '文件上传'},
    'uploadFile(word)': {'component': 'Upload', 'label': '附件上传（Word）'},
    'table': {'component': 'Table', 'label': '表格'},
    'textJson': {'component': 'Input.TextArea', 'label': 'Json编辑器'},
    'selectorIcon': {'component': 'IconPicker', 'label': '图标选择器'},
    'richTextLocal': {'component': 'RichTextEditor', 'label': '富文本（本地图片）'},
    'richTextURL': {'component': 'RichTextEditor', 'label': '富文本（网络图片）'},
}

# 中文标签 → 英文编码 反向映射
LABEL_TO_CODE = {v['label']: k for k, v in INPUT_TYPE_MAPPING.items()}

# 卡片宽度映射
CARD_WIDTH_MAPPING = {
    '四分之一行': {'span': 6, 'percent': '25%'},
    '三分之一行': {'span': 8, 'percent': '33%'},
    '半行': {'span': 12, 'percent': '50%'},
    '整行': {'span': 24, 'percent': '100%'},
}

# 控制状态映射
CONTROL_MAPPING = {
    '读写': {'editable': True, 'disabled': False, 'visible': True},
    '只读': {'editable': False, 'disabled': True, 'visible': True},
    '隐藏': {'editable': False, 'disabled': False, 'visible': False},
    '禁用': {'editable': False, 'disabled': True, 'visible': True},
}


def parse_bool(value):
    """解析布尔值"""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ['是', 'yes', 'true', '1']
    return False


def parse_int(value):
    """解析整数"""
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def parse_input_params(value):
    """解析输入参数"""
    if value is None:
        return None
    value = str(value).strip()
    if not value:
        return None

    # JSON 格式（dict 引用）
    if value.startswith('{'):
        try:
            parsed = json.loads(value)
            if parsed.get('type') == 'dict':
                return {
                    'type': 'dict',
                    'dict_code': parsed.get('param', {}).get('code'),
                    'raw': value
                }
        except json.JSONDecodeError:
            pass

    # 逗号分隔（字面枚举）
    if ',' in value or '，' in value:
        options = [v.strip() for v in value.replace('，', ',').split(',') if v.strip()]
        return {
            'type': 'enum',
            'options': [{'label': opt, 'value': opt} for opt in options],
            'raw': value
        }

    # 其他（字面量）
    return {'type': 'literal', 'value': value, 'raw': value}


def get_column_headers(ws):
    """获取列头映射（第 2 行是实际列名）"""
    headers = {}

    # 读取第 2 行（实际列名）
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=col).value
        if cell_value:
            col_name = str(cell_value).strip()
            if col_name in COLUMN_MAPPING:
                headers[col] = COLUMN_MAPPING[col_name]

    return headers


def parse_excel(file_path):
    """解析 Excel 文件"""
    print(f"正在解析: {file_path}")

    wb = load_workbook(file_path, data_only=True)

    # 获取主表（第一个工作表）
    ws = wb.worksheets[0]
    print(f"工作表: {ws.title}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")

    # 获取列头映射
    headers = get_column_headers(ws)
    print(f"识别到 {len(headers)} 个有效列")

    if not headers:
        print("错误：无法识别列头")
        return None

    # 解析数据行
    forms = {}  # 按表单编码分组
    current_row = 3  # 假设数据从第3行开始（跳过双层表头）

    for row in range(current_row, ws.max_row + 1):
        # 检查是否是空行
        first_cell = ws.cell(row=row, column=1).value
        if first_cell is None or str(first_cell).strip() == '':
            continue

        # 提取字段数据
        field = {}
        for col, field_name in headers.items():
            cell_value = ws.cell(row=row, column=col).value
            field[field_name] = cell_value

        # 获取表单编码
        form_code = field.get('form_code')
        if not form_code:
            continue

        form_code = str(form_code).strip()

        # 初始化表单
        if form_code not in forms:
            forms[form_code] = {
                'form_code': form_code,
                'form_name': str(field.get('form_name', '')).strip(),
                'app_name': str(field.get('app_name', '')).strip(),
                'filter_condition': str(field.get('filter_condition', '')).strip(),
                'fields': [],
                'groups': {}
            }

        # 处理字段数据
        processed_field = {
            'field_name': str(field.get('field_name', '')).strip(),
            'field_code': str(field.get('field_code', '')).strip(),
            'is_primary_key': parse_bool(field.get('is_primary_key')),
            'is_virtual': parse_bool(field.get('is_virtual')),
            'field_type': str(field.get('field_type', '')).strip() if field.get('field_type') else None,
            'field_length': parse_int(field.get('field_length')),
            'field_precision': parse_int(field.get('field_precision')),
            'default_value': str(field.get('default_value', '')).strip() if field.get('default_value') else None,

            'input_type': str(field.get('input_type', '')).strip() if field.get('input_type') else None,
            'input_params': parse_input_params(field.get('input_params')),
            'is_required': parse_bool(field.get('is_required')),
            'is_unique': parse_bool(field.get('is_unique')),
            'placeholder': str(field.get('placeholder', '')).strip() if field.get('placeholder') else None,
            'remark': str(field.get('remark', '')).strip() if field.get('remark') else None,

            'card_group': str(field.get('card_group', '')).strip() if field.get('card_group') else '基本信息',
            'card_sort': parse_int(field.get('card_sort')),
            'card_width': str(field.get('card_width', '')).strip() if field.get('card_width') else '半行',
            'add_control': str(field.get('add_control', '')).strip() if field.get('add_control') else '读写',
            'update_control': str(field.get('update_control', '')).strip() if field.get('update_control') else '读写',
            'detail_control': str(field.get('detail_control', '')).strip() if field.get('detail_control') else '读写',

            'list_width': parse_int(field.get('list_width')),
            'list_control': str(field.get('list_control', '')).strip() if field.get('list_control') else '显示',
            'list_sort': parse_int(field.get('list_sort')),
            'list_formatter': str(field.get('list_formatter', '')).strip() if field.get('list_formatter') else None,

            'is_filter': parse_bool(field.get('is_filter')),
            'filter_mode': str(field.get('filter_mode', '')).strip() if field.get('filter_mode') else None,
            'filter_default': str(field.get('filter_default', '')).strip() if field.get('filter_default') else None,
            'filter_placeholder': str(field.get('filter_placeholder', '')).strip() if field.get('filter_placeholder') else None,
            'source_system': str(field.get('source_system', '')).strip() if field.get('source_system') else None,
        }

        # 映射组件信息（处理中文标签）
        input_type = processed_field['input_type']
        # 先尝试英文编码，再尝试中文标签
        if input_type in INPUT_TYPE_MAPPING:
            processed_field['input_type_code'] = input_type
            processed_field['input_component'] = INPUT_TYPE_MAPPING[input_type]['component']
            processed_field['input_type_label'] = INPUT_TYPE_MAPPING[input_type]['label']
        elif input_type in LABEL_TO_CODE:
            # 中文标签 → 英文编码
            code = LABEL_TO_CODE[input_type]
            processed_field['input_type_code'] = code
            processed_field['input_component'] = INPUT_TYPE_MAPPING[code]['component']
            processed_field['input_type_label'] = input_type
        else:
            processed_field['input_type_code'] = input_type
            processed_field['input_component'] = 'Input'
            processed_field['input_type_label'] = input_type

        # 映射卡片宽度
        card_width = processed_field['card_width']
        if card_width in CARD_WIDTH_MAPPING:
            processed_field['card_width_span'] = CARD_WIDTH_MAPPING[card_width]['span']
        else:
            processed_field['card_width_span'] = 12  # 默认半行

        # 添加到表单
        forms[form_code]['fields'].append(processed_field)

        # 按卡片分组
        card_group = processed_field['card_group']
        if card_group not in forms[form_code]['groups']:
            forms[form_code]['groups'][card_group] = []
        forms[form_code]['groups'][card_group].append(processed_field)

    # 转换分组为数组格式
    result_forms = []
    for form_code, form_data in forms.items():
        groups = []
        for group_name, group_fields in form_data['groups'].items():
            groups.append({
                'group_name': group_name,
                'group_sort': group_fields[0].get('card_sort', 99) if group_fields else 99,
                'default_span': 12,
                'fields': group_fields
            })

        # 按排序字段排序
        groups.sort(key=lambda x: x['group_sort'] if x['group_sort'] else 99)

        # 生成表格列定义
        table_columns = []
        for field in form_data['fields']:
            if field.get('list_control') != '隐藏':
                table_columns.append({
                    'title': field['field_name'],
                    'dataIndex': field['field_code'],
                    'key': field['field_code'],
                    'width': field.get('list_width') or 120,
                    'valueType': field.get('input_component', 'text').lower(),
                })

        result_forms.append({
            'form_code': form_code,
            'form_name': form_data['form_name'],
            'app_name': form_data['app_name'],
            'filter_condition': form_data['filter_condition'],
            'field_count': len(form_data['fields']),
            'groups': groups,
            'table_columns': table_columns,
            'fields': form_data['fields']
        })

    return result_forms


def main():
    """主函数"""
    import argparse
    parser = argparse.ArgumentParser(description='解析 Excel 表单模板')
    parser.add_argument('excel_file', help='Excel 文件路径')
    parser.add_argument('-o', '--output-dir', default=None, help='输出目录（默认: data/forms）')
    args = parser.parse_args()

    # 文件路径
    excel_file = args.excel_file
    if not os.path.exists(excel_file):
        print(f"错误：文件不存在: {excel_file}")
        return

    output_dir = args.output_dir or os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'forms')

    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)

    # 解析 Excel
    forms = parse_excel(excel_file)

    if not forms:
        print("解析失败")
        return

    # 输出统计
    print(f"\n解析完成:")
    print(f"- 表单数量: {len(forms)}")
    print(f"- 总字段数: {sum(f['field_count'] for f in forms)}")

    # 保存为 JSON
    output_file = os.path.join(output_dir, 'forms.json')
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({
            'meta': {
                'generated_at': datetime.now().isoformat(),
                'source_file': os.path.basename(excel_file),
                'total_forms': len(forms),
                'total_fields': sum(f['field_count'] for f in forms)
            },
            'forms': forms
        }, f, ensure_ascii=False, indent=2)

    print(f"\n已保存到: {output_file}")

    # 打印每个表单的统计
    print("\n表单列表:")
    for form in forms:
        print(f"  - {form['form_code']}: {form['form_name']} ({form['field_count']} 字段)")


if __name__ == '__main__':
    main()
