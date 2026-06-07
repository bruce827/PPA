#!/usr/bin/env python3
"""探索 Excel 文件结构
用法: python3 explore_excel.py <excel文件路径>
"""

import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

if len(sys.argv) < 2:
    print("用法: python3 explore_excel.py <excel文件路径>")
    sys.exit(1)

file_path = sys.argv[1]

wb = load_workbook(file_path, data_only=True)
ws = wb.worksheets[0]

print(f"工作表: {ws.title}")
print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
print()

# 打印前 5 行的所有列
print("前 5 行内容:")
for row in range(1, 6):
    print(f"\n第 {row} 行:")
    for col in range(1, min(ws.max_column + 1, 20)):  # 只显示前 20 列
        cell_value = ws.cell(row=row, column=col).value
        if cell_value:
            col_letter = get_column_letter(col)
            print(f"  {col_letter}({col}): {repr(cell_value)}")
